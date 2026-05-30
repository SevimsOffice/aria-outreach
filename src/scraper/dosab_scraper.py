"""
DOSAB (Demirtaş OSB) scraper — dosab.org.tr

DOSAB provides a downloadable Excel file with all 573 member companies.
This scraper:
  1. Fetches the list page to find the current Excel download URL (dynamic — URL changes with updates)
  2. Downloads the Excel file
  3. Parses columns using keyword detection (handles Turkish headers)
"""

import io
import logging
import re
from urllib.parse import urljoin

import openpyxl
import requests

from .base_scraper import (
    get_session, fetch_page, extract_domain,
    normalize_phone, clean_company_name, HEADERS
)

logger = logging.getLogger(__name__)

BASE_URL   = "https://www.dosab.org.tr"
LIST_PAGE  = f"{BASE_URL}/Detay/730/Firma-Listesi"
OSB_NAME   = "DOSAB"

# Keyword → field mapping (lowercase match against header cell values)
COLUMN_KEYWORDS: dict[str, str] = {
    "firma adı":     "Company_Name",
    "firma adi":     "Company_Name",
    "şirket adı":    "Company_Name",
    "sirket adi":    "Company_Name",
    "unvan":         "Company_Name",
    "ad/unvan":      "Company_Name",
    "sektör":        "Sector",
    "sektor":        "Sector",
    "faaliyet":      "Sector",
    "alan":          "Sector",
    "adres":         "Address",
    "telefon":       "Phone",
    "tel":           "Phone",
    "e-posta":       "Email",
    "eposta":        "Email",
    "e-mail":        "Email",
    "email":         "Email",
    "mail":          "Email",
    "web":           "Domain",
    "website":       "Domain",
    "internet":      "Domain",
    "site":          "Domain",
}


def scrape_dosab() -> list[dict]:
    """
    Download and parse DOSAB member Excel file.
    Returns list of company dicts (Company_Name, Sector, Domain, Phone, Address, Email, OSB).
    """
    session = get_session()

    # Step 1: Find Excel download URL
    excel_url = _find_excel_url(session)
    if not excel_url:
        logger.error("DOSAB: could not find Excel download link on list page")
        return []

    # Step 2: Download Excel
    logger.info(f"DOSAB: downloading Excel from {excel_url}")
    try:
        resp = session.get(excel_url, timeout=45)
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"DOSAB: Excel download failed: {e}")
        return []

    # Step 3: Parse
    try:
        companies = _parse_excel(resp.content)
        logger.info(f"DOSAB: parsed {len(companies)} companies")
        return companies
    except Exception as e:
        logger.error(f"DOSAB: Excel parsing failed: {e}", exc_info=True)
        return []


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _find_excel_url(session) -> str:
    """Scrape the DOSAB list page to find the current Excel download URL."""
    soup = fetch_page(LIST_PAGE, session)
    if not soup:
        return ""

    for a in soup.find_all("a", href=True):
        href = a["href"]
        if ".xlsx" in href.lower() or ".xls" in href.lower():
            if href.startswith("http"):
                return href
            return urljoin(BASE_URL, href)
    return ""


def _parse_excel(content: bytes) -> list[dict]:
    """Parse Excel bytes → list of company dicts with dynamic column detection."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # ── Find header row (first row that has at least 2 non-empty cells) ──
    header_row_idx = None
    header_values: list[str] = []
    for i, row in enumerate(ws.iter_rows(max_row=10), start=1):
        vals = [str(cell.value or "").strip() for cell in row]
        non_empty = [v for v in vals if v]
        if len(non_empty) >= 2:
            header_row_idx = i
            header_values = vals
            break

    if header_row_idx is None:
        logger.error("DOSAB Excel: no header row found in first 10 rows")
        return []

    logger.debug(f"DOSAB Excel headers: {header_values}")

    # ── Map column index → field name ──
    col_map: dict[int, str] = {}
    for j, raw_header in enumerate(header_values):
        normalized = raw_header.lower().strip()
        for keyword, field in COLUMN_KEYWORDS.items():
            if keyword in normalized:
                if field not in col_map.values():   # First match wins per field
                    col_map[j] = field
                break

    if not col_map:
        # Fallback: treat column 0 as company name, column 1 as address
        logger.warning("DOSAB Excel: no matching headers found; using positional fallback")
        col_map = {0: "Company_Name"}
        if len(header_values) > 1:
            col_map[1] = "Address"
        if len(header_values) > 2:
            col_map[2] = "Phone"

    logger.debug(f"DOSAB column mapping: {col_map}")

    # ── Parse data rows ──
    companies: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        vals = [str(cell.value or "").strip() for cell in row]
        if not any(vals):
            continue    # Skip blank rows

        record: dict = {
            "Company_Name": "",
            "Sector":       "",
            "Domain":       "",
            "Phone":        "",
            "Address":      "",
            "Email":        "",
            "OSB":          OSB_NAME,
            "City":         "Bursa",
        }

        for col_idx, field in col_map.items():
            if col_idx >= len(vals):
                continue
            val = vals[col_idx]
            if not val:
                continue
            if field == "Domain":
                record["Domain"] = extract_domain(val)
            elif field == "Phone":
                record["Phone"] = normalize_phone(val)
            elif field == "Company_Name":
                record["Company_Name"] = clean_company_name(val)
            else:
                record[field] = val

        # Also scan all columns for email via regex (some sheets embed emails in address columns)
        if not record["Email"]:
            full_row_text = " ".join(vals)
            m = re.search(
                r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
                full_row_text,
            )
            if m:
                record["Email"] = m.group().lower()

        if record["Company_Name"] and len(record["Company_Name"]) > 2:
            companies.append(record)

    return companies
